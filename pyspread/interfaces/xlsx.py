#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Copyright Martin Manns
# Distributed under the terms of the GNU General Public License

# --------------------------------------------------------------------
# pyspread is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# pyspread is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with pyspread.  If not, see <http://www.gnu.org/licenses/>.
# --------------------------------------------------------------------

"""

This file contains interfaces to the xlsx file format.

XlsReader and XlsWriter classed are structured into the following sections:
 * shape
 * code
 * attributes
 * row_heights
 * col_widths
 * macros


**Provides**

 * :func:`wxcolor2rgb`
 * :dict:`wx2qt_fontweights`
 * :dict:`wx2qt_fontstyles`
 * :class:`XlsReader`
 * :class:`XlsWriter`

"""

# from builtins import str, map, object

from collections import defaultdict
import logging
from typing import BinaryIO

try:
    import pycel
    from pycel import excelformula
except ImportError:
    pycel = None

try:
    from openpyxl import load_workbook, worksheet
    from openpyxl.workbook import Workbook
    from openpyxl.cell.cell import (Cell, TYPE_STRING, TYPE_FORMULA,
                                    TYPE_NUMERIC, TYPE_BOOL, TYPE_NULL,
                                    TYPE_INLINE, TYPE_ERROR,
                                    TYPE_FORMULA_CACHE_STRING)
    from openpyxl.utils.cell import column_index_from_string
except ImportError:
    openpyxl = None
    worksheet = None
    Cell = None
    Workbook = None

try:
    from pyspread.lib.attrdict import AttrDict
    from pyspread.lib.selection import Selection
    from pyspread.lib.parsers import spreadsheet_formula_to_code
    from pyspread.model.model import CellAttribute, CodeArray
except ImportError:
    from lib.attrdict import AttrDict
    from lib.selection import Selection
    from lib.parsers import spreadsheet_formula_to_code
    from model.model import CellAttribute, CodeArray


TYPE_DATE = 'd'

BORDERWIDTH_XLSX2PYSU = {"hair": 0, "thin": 1, "medium": 4, "thick": 8}


class XlsxReader:
    """Reads xlsx file from OpenPyXL into a code_array"""

    def __init__(self, xlsx_file: BinaryIO, code_array: CodeArray):
        """
        :param xlsx_file: The xlsx file to be read
        :param code_array: Target code_array

        """

        self.code_array = code_array
        self.wb = load_workbook(xlsx_file)

    def __iter__(self):
        """Iterates over self.xlsx_file, replacing everything in code_array"""

        sheet_attrs = defaultdict(list)

        self._xlsx2shape(self.wb)

        for worksheet_name in self.wb.sheetnames:
            worksheet = self.wb[worksheet_name]
            table = self.wb.sheetnames.index(worksheet_name)

            # Row widths
            self._xlsx2row_heights(worksheet, table)

            # Column widths
            self._xlsx2col_widths(worksheet, table)

            # Merged cells
            self._xlsx2merge_areas(worksheet, table, sheet_attrs)
            self.append_sheet_attributes(table, sheet_attrs)
            sheet_attrs.clear()

            # Cells
            for row_idx, row in enumerate(worksheet.iter_rows()):
                for cell in row:
                    key = row_idx, cell.column-1, table

                    # Code
                    self._xlsx2code(key, cell)

                    # Cell attributes
                    self._xlsx2attributes(key, cell, sheet_attrs)

                    # cell.fill, cell.alignment, cell.border, cell.fill,
                    # cell.font, cell.has_style, cell.hyperlink,
                    # cell.is_date, cell.number_format, cell.protection

                    yield key

            self.append_sheet_attributes(table, sheet_attrs)
            sheet_attrs.clear()

    def append_sheet_attributes(self, table: int, sheet_attrs: dict):
        """Creates selections from attributes in  sheet_attrs and appends them

        :param table: Current table
        :param sheet_attrs: Dict mapping (attr_name, attr values) to key list

        """

        for sheet_attr in sheet_attrs:
            selection = Selection([], [], [], [], sheet_attrs[sheet_attr])
            attr_dict = AttrDict([sheet_attr])
            cell_attr = CellAttribute(selection, table, attr_dict)
            self.code_array.cell_attributes.append(cell_attr)

    @staticmethod
    def xlsx_rgba2rgb(rgb: str) -> (int, int, int):
        """Converts rgba string from xlsx to rgb tuple

        :param rgb: rgba string from xlsx

        """

        if rgb[:2] == "00":
            # Transparent cell
            raise ValueError("Color is transparent")

        r = int(rgb[2:4], 16)
        g = int(rgb[4:6], 16)
        b = int(rgb[6:8], 16)

        return r, g, b

    def _xlsx2shape(self, wb: Workbook):
        """Updates shape in code_array

        :param wb: openpyxl workbook

        """

        sheetnames = self.wb.sheetnames

        max_row = 1
        max_column = 1

        for worksheet_name in sheetnames:
            worksheet = self.wb[worksheet_name]
            if max_row < worksheet.max_row:
                max_row = worksheet.max_row
            if max_column < worksheet.max_column:
                max_column = worksheet.max_column

        self.code_array.macros += f"_sheetnames = {sheetnames}"

        shape = max_row, max_column, len(sheetnames)
        self.code_array.shape = shape

    def _xlsx2row_heights(self, worksheet: worksheet, table: int):
        """Updates row_heights in code_array

        :param worksheet: openpyxl worksheet
        :param table: pyspread table number

        """

        for row in worksheet.row_dimensions:
            height = worksheet.row_dimensions[row].height / 12.8 * 30
            key = row - 1, table
            self.code_array.row_heights[key] = height

    def _xlsx2col_widths(self, worksheet: worksheet, table: int):
        """Updates col_widths in code_array

        :param worksheet: openpyxl worksheet
        :param table: pyspread table number

        """

        for column in worksheet.column_dimensions:
            width = worksheet.column_dimensions[column].width / 11.26 * 100
            key = column_index_from_string(column) - 1, table
            self.code_array.col_widths[key] = width

    @staticmethod
    def _xlsx2merge_areas(worksheet, table, sheet_attrs):
        """Updates merge_areas in sheet_attrs

        :param worksheet: openpyxl worksheet
        :param table: pyspread table number
        :param sheet_attrs: Dict mapping (attr_name, attr values) to key list

        """

        for merged_range in worksheet.merged_cells.ranges:
            area = (merged_range.min_row-1, merged_range.min_col-1,
                    merged_range.max_row-1, merged_range.max_col-1)
            skey = merged_range.min_row-1, merged_range.min_col-1, table
            sheet_attrs[("merge_area", area)].append(skey)

    def _xlsx2code(self, key: (int, int, int),  cell: Cell):
        """Updates code in code_array

        :param key: Pyspread cell key
        :param cell: Cell object from openpyxl

        """

        if pycel is None or cell.data_type in (TYPE_NUMERIC, TYPE_STRING,
                                               TYPE_INLINE, TYPE_DATE,
                                               TYPE_FORMULA_CACHE_STRING):
            code = repr(cell.value)
        elif cell.data_type == TYPE_FORMULA:
            # Convert formula via pycel
            logging.debug(f"Cell {key} is coverted via pycel")
            code = spreadsheet_formula_to_code(cell.value)
        elif cell.data_type == TYPE_BOOL:
            code = "True" if cell.value else "False"
        elif cell.data_type == TYPE_NULL:
            code = None
        elif cell.data_type == TYPE_ERROR:
            code = Exception(str(cell.value))
        else:
            code = repr(cell.value)
            msg = f"Excel data type {cell.data_type} of cell {cell} unknown."
            logging.error(msg)

        logging.debug(f"Cell {key} set to {code}")

        self.code_array.dict_grid[key] = code

    def _xlsx2attributes(self, key: (int, int, int), cell: Cell, sheet_attrs):
        """Updates attributes in code_array

        :param key: Pyspread cell key
        :param cell: Cell object from openpyxl
        :param sheet_attrs: Sheet attribute dict

        """

        skey = key[0], key[1]  # sheet key
        skey_above = key[0] - 1, key[1]
        skey_left = key[0], key[1] - 1

        # Cell background color
        logging.debug("Cell background color")
        try:
            bg_rgb = self.xlsx_rgba2rgb(cell.fill.bgColor.rgb)
            sheet_attrs[("bgcolor", bg_rgb)].append(skey)
        except (ValueError, TypeError):
            pass

        # Text color
        logging.debug("Text color")
        try:
            text_rgb = self.xlsx_rgba2rgb(cell.font.color.rgb)
            sheet_attrs[("textcolor", text_rgb)].append(skey)
        except (AttributeError, ValueError, TypeError):
            pass

        # Text font
        logging.debug("Text font")
        try:
            sheet_attrs[("textfont", cell.font.name)].append(skey)
        except ValueError:
            pass

        # Text size
        logging.debug("Text size")
        try:
            sheet_attrs[("pointsize", cell.font.size)].append(skey)
        except ValueError:
            pass

        # Text weight
        logging.debug("Text weight")
        try:
            if cell.font.bold:
                sheet_attrs[("fontweight", 75)].append(skey)
        except ValueError:
            pass

        # Text style
        logging.debug("Text style")
        try:
            if cell.font.italic:
                sheet_attrs[("fontstyle", 1)].append(skey)
        except ValueError:
            pass

        # Text underline
        logging.debug("Text underline")
        try:
            if cell.font.underline:
                sheet_attrs[("underline", True)].append(skey)
        except ValueError:
            pass

        # Text strikethrough
        logging.debug("Text strikethrough")
        try:
            if cell.font.strike:
                sheet_attrs[("strikethrough", True)].append(skey)
        except ValueError:
            pass

        # Cell alignment
        logging.debug("Cell alignment")

        if cell.alignment.horizontal == "left":
            sheet_attrs[("justification", "justify_left")].append(skey)
        elif cell.alignment.horizontal == "center":
            sheet_attrs[("justification", "justify_center")].append(skey)
        elif cell.alignment.horizontal == "justify":
            sheet_attrs[("justification", "justify_fill")].append(skey)
        elif cell.alignment.horizontal == "right":
            sheet_attrs[("justification", "justify_right")].append(skey)

        if cell.alignment.vertical == "top":
            sheet_attrs[("vertical_align", "align_top")].append(skey)
        elif cell.alignment.vertical == "center":
            sheet_attrs[("vertical_align", "align_center")].append(skey)
        elif cell.alignment.vertical == "bottom":
            sheet_attrs[("vertical_align", "align_bottom")].append(skey)

        # Cell borders
        logging.debug("Cell borders")

        if cell.border.bottom.style is not None:
            width = BORDERWIDTH_XLSX2PYSU[cell.border.bottom.style]
            sheet_attrs[("borderwidth_bottom", width)].append(skey)

        try:
            rgb = cell.border.bottom.color.rgb
            color = self.xlsx_rgba2rgb(rgb)
            sheet_attrs[("bordercolor_bottom", color)].append(skey)
        except (AttributeError, TypeError):
            if cell.border.bottom.style is not None and width is not None:
                color = 0, 0, 0
                sheet_attrs[("bordercolor_bottom", color)].append(skey)

        if cell.border.right.style is not None:
            width = BORDERWIDTH_XLSX2PYSU[cell.border.right.style]
            sheet_attrs[("borderwidth_right", width)].append(skey)

        try:
            rgb = cell.border.right.color.rgb
            color = self.xlsx_rgba2rgb(rgb)
            sheet_attrs[("bordercolor_right", color)].append(skey)
        except (AttributeError, TypeError):
            if cell.border.right.style is not None and width is not None:
                color = 0, 0, 0
                sheet_attrs[("bordercolor_right",
                             color)].append(skey)

        if cell.border.top.style is not None:
            width = BORDERWIDTH_XLSX2PYSU[cell.border.top.style]
            sheet_attrs[("borderwidth_bottom",
                         width)].append(skey_above)

        try:
            rgb = cell.border.top.color.rgb
            color = self.xlsx_rgba2rgb(rgb)
            sheet_attrs[("bordercolor_bottom",
                         color)].append(skey_above)
        except (AttributeError, TypeError):
            if cell.border.top.style is not None and width is not None:
                color = 0, 0, 0
                sheet_attrs[("bordercolor_bottom",
                             color)].append(skey_above)

        if cell.border.left.style is not None:
            width = BORDERWIDTH_XLSX2PYSU[cell.border.left.style]
            sheet_attrs[("borderwidth_right",
                         width)].append(skey_left)

        try:
            rgb = cell.border.left.color.rgb
            color = self.xlsx_rgba2rgb(rgb)
            sheet_attrs[("bordercolor_right",
                         color)].append(skey_left)
        except (AttributeError, TypeError):
            if cell.border.left.style is not None and width is not None:
                color = 0, 0, 0
                sheet_attrs[("bordercolor_right",
                             color)].append(skey_left)
